# ----- Librerias ------

import tkinter as tk
from tkinter import messagebox, simpledialog
import pandas as pd
import cv2  # Para la captura de QR
from pyzbar.pyzbar import decode  # Decodificar QR
import os
import requests
import json
from dotenv import load_dotenv
import urllib3
import re
import numpy as np
import threading
import queue
from openpyxl import load_workbook

# ---------- Configuraciones -------------

# Deshabilitar las advertencias de SSL
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Cargar las variables del archivo .env
load_dotenv()

# Configuración de la API de GLPI
GLPI_URL = os.getenv("GLPI_URL")
USER_TOKEN = os.getenv("USER_TOKEN")
APP_TOKEN = os.getenv("APP_TOKEN")
PATH_EXCEL_ACTIVOS = os.getenv("PATH_EXCEL_ACTIVOS")
PATH_EXCEL_CONSUMIBLES = os.getenv("PATH_EXCEL_CONSUMIBLES")
IP_CAM_URL = os.getenv("IP_CAM_URL")

# Ruta del archivo Excel
ruta_excel = PATH_EXCEL_ACTIVOS
ruta_excel_consumibles = PATH_EXCEL_CONSUMIBLES

# Crear archivo Excel si no existe
def crear_archivo_excel(ruta, columnas):
    if not os.path.exists(ruta):
        df = pd.DataFrame(columns=columnas)
        df.to_excel(ruta, index=False)

crear_archivo_excel(ruta_excel, ["Asset Type", "Name", "Location", "Manufacturer", "Model", "Serial Number", 
                                 "Inventory Number", "Comments", "Technician in Charge", "Group in Charge", "Status", "Specific Fields (Dynamic Column)"])
crear_archivo_excel(ruta_excel_consumibles, ["Name", "Inventory/Asset Number", "Location", "Stock Target"])

class GLPIApp:
    # ---- Configs. iniciales ----
    
    def __init__(self, root):
        self.root = root
        self.root.title("GLPI Asset Automator")
        self.create_widgets()
        
    def create_widgets(self):
        frame = tk.Frame(self.root)
        frame.pack(padx=10, pady=10)

        # Laptops
        tk.Label(frame, text="----- Laptops -----").grid(row=0, column=0, columnspan=2)
        tk.Button(frame, text="Escanear QR y registrar laptop (Dell/Mac)", command=lambda: self.manejar_qr_laptop("Register")).grid(row=1, column=0)
        tk.Button(frame, text="Entregar laptop a un usuario", command=self.entregar_laptop).grid(row=1, column=1)

        # Monitores
        tk.Label(frame, text="----- Monitores -----").grid(row=2, column=0, columnspan=2)
        tk.Button(frame, text="Escanear QR y registrar monitores", command=self.manejar_qr_monitor).grid(row=3, column=0)
        tk.Button(frame, text="Entregar monitor a un usuario", command=self.entregar_monitor).grid(row=3, column=1)

        # Consumibles
        tk.Label(frame, text="----- Consumibles -----").grid(row=4, column=0, columnspan=2)
        tk.Button(frame, text="Agregar consumible", command=self.agregar_consumible).grid(row=5, column=0)
        tk.Button(frame, text="Quitar consumible", command=self.quitar_consumible).grid(row=5, column=1)

        # Excel
        tk.Label(frame, text="----- Excel -----").grid(row=6, column=0, columnspan=2)
        tk.Button(frame, text="Registrar la última fila del Excel en GLPI", command=self.registrar_ultima_fila).grid(row=7, column=0)
        tk.Button(frame, text="Registrar un activo por nombre", command=self.registrar_por_nombre).grid(row=7, column=1)
        tk.Button(frame, text="Registrar todos los activos de Excel en GLPI", command=lambda: self.procesar_archivo_excel(ruta_excel)).grid(row=8, column=0, columnspan=2)
        tk.Button(frame, text="Salir", command=self.root.quit).grid(row=9, column=0, columnspan=2)
        
    # ------- Funciones -----------
    # ----- Excel y configs. ------
    
    def obtener_token_sesion(self):
        def run_request():
            headers = {
                "Authorization": f"user_token {USER_TOKEN}",
                "App-Token": APP_TOKEN,
            }
            response = requests.get(f"{GLPI_URL}/initSession", headers=headers, verify=False)
            if response.status_code == 200:
                self.session_token = response.json().get("session_token")
            else:
                messagebox.showerror("Error", f"Error al iniciar sesión: {response.status_code}")

        request_thread = threading.Thread(target=run_request)
        request_thread.start()
        
    def verificar_existencia_asset(self, session_token, serial_number, asset_type="Computer"):
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        # Definir los endpoints válidos para diferentes tipos de activos
        endpoints = {
            "Computer": "Computer",
            "Monitor": "Monitor",
            "Network Equipment": "NetworkEquipment",
            "Consumables": "ConsumableItem",
        }

        endpoint = endpoints.get(asset_type, "Computer")

        params = {"searchText": serial_number, "range": "0-999"}
        response = requests.get(f"{GLPI_URL}/search/{endpoint}", headers=headers, params=params, verify=False)

        if response.status_code == 200:
            assets = response.json().get("data", [])
            for asset in assets:
                if asset.get('5', '').strip() == serial_number:
                    messagebox.showinfo("Información", f"El activo con número de serie '{serial_number}' ya existe en GLPI con ID: {asset.get('1')}")
                    return True
        else:
            messagebox.showerror("Error", f"Error al verificar existencia del activo: {response.status_code}")
            try:
                messagebox.showerror("Error", response.json())
            except json.JSONDecodeError:
                messagebox.showerror("Error", response.text)
        
        return False  
    
    def verificar_existencia_en_excel(self, serial_number):
        wb = load_workbook(ruta_excel)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):
            if row[5] == serial_number:  # Asumiendo que la columna "Serial Number" es la sexta
                messagebox.showinfo("Información", f"El activo con número de serie '{serial_number}' ya existe en el Excel.")
                return True
        return False

    def agregar_a_excel(self, asset_data):
        try:
            df = pd.read_excel(ruta_excel)

            if self.verificar_existencia_en_excel(asset_data["Serial Number"]):
                messagebox.showinfo("Información", f"El activo con serial '{asset_data['Serial Number']}' ya está registrado en el Excel. No se agregará.")
                return

            nuevo_registro = pd.DataFrame([asset_data])  # Convertir el asset_data a un DataFrame de una fila
            df = pd.concat([df, nuevo_registro], ignore_index=True)  # Agregar la nueva fila al DataFrame existente
            df.to_excel(ruta_excel, index=False)  # Guardar el DataFrame actualizado en el Excel
            messagebox.showinfo("Éxito", "Datos registrados exitosamente en el Excel.")

            # Guardar la plantilla en un archivo .txt
            nombre_archivo_txt = f"C:/Users/sebastian.salgado/Desktop/GLPI-Asset-Automator/Templates/{asset_data['Name']}.txt"
            self.guardar_plantilla_txt(asset_data, nombre_archivo_txt)

            # Preguntar si se desea registrar en GLPI
            registrar_glpi = simpledialog.askstring("Registrar en GLPI", "¿Deseas registrar este activo en GLPI? (sí/no):").strip().lower()
            if registrar_glpi == "sí" or registrar_glpi == "si":
                self.registrar_ultima_fila()
            else:
                messagebox.showinfo("Información", "El activo no fue registrado en GLPI.")
        except Exception as e:
            messagebox.showerror("Error", f"Error al guardar los datos: {e}")

    def registrar_asset(self, session_token, asset_data, asset_type):
        if self.verificar_existencia_asset( session_token, asset_data["serial"]):
            messagebox.showinfo("Información", f"El activo con número de serie {asset_data['serial']} ya existe en GLPI. No se realizará el registro.")
            return

        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        endpoint = {
            "Computer": "/Computer",
            "Monitor": "/Monitor",
            "Network Equipment": "/NetworkEquipment",
            "Consumables": "/ConsumableItem",
        }.get(asset_type, "/Computer")

        # Limpiar datos antes de enviarlos
        asset_data_clean = self.limpiar_asset_data(asset_data)

        # Crear la estructura correcta para la API de GLPI
        asset_data_array = {"input": [asset_data_clean]}
        
        response = requests.post(f"{GLPI_URL}{endpoint}", headers=headers, data=json.dumps(asset_data_array), verify=False)
        
        if response.status_code == 201:
            messagebox.showinfo("Éxito", f"Asset registrado exitosamente: {asset_data_clean['name']}")
        else:
            messagebox.showerror("Error", f"Error al registrar asset: {response.status_code}")
            try:
                messagebox.showerror("Error", response.json())
            except json.JSONDecodeError:
                messagebox.showerror("Error", response.text)

    def registrar_ultima_fila(self):
        df = pd.read_excel(ruta_excel)
        if df.empty:
            messagebox.showerror("Error", "El archivo Excel está vacío.")
            return

        last_row = df.iloc[-1].to_dict()
        #messagebox.showinfo("Información", f"Última fila encontrada: {last_row}")

        # Reemplazar NaN con valores vacíos
        last_row = {key: ("" if pd.isna(value) else value) for key, value in last_row.items()}

        session_token = self.obtener_token_sesion()
        if not session_token:
            messagebox.showerror("Error", "No se pudo obtener el token de sesión.")
            return

        if "Name" not in last_row or "Asset Type" not in last_row:
            messagebox.showerror("Error", "La última fila no contiene las columnas esperadas.")
            return

        location_id = self.obtener_location_id(session_token, last_row["Location"])
        if not location_id:
            messagebox.showerror("Error", f"No se pudo encontrar la ubicación: {last_row['Location']}")
            return

        manufacturer_id = self.obtener_manufacturer_id(session_token, last_row["Manufacturer"])
        if not manufacturer_id:
            messagebox.showerror("Error", f"No se pudo encontrar el fabricante: {last_row['Manufacturer']}")
            return

        if location_id is None or manufacturer_id is None:
            messagebox.showerror("Error", f"No se pudo encontrar la ubicación o el fabricante para el activo '{last_row['Name']}'")
            return

        # Preparar los datos para el registro en GLPI
        asset_data = {
            "name": last_row["Name"].strip(),
            "locations_id": location_id,
            "manufacturers_id": manufacturer_id,
            "serial": last_row["Serial Number"].strip(),
            "comments": last_row["Comments"].strip() if last_row["Comments"] else "N/A",
        }

        messagebox.showinfo("Información", f"Registrando asset: {asset_data}")

        self.registrar_asset(session_token, asset_data, last_row["Asset Type"])

    def registrar_por_nombre(self):
        df = pd.read_excel(ruta_excel)
        if df.empty:
            messagebox.showerror("Error", "El archivo Excel está vacío.")
            return

        nombre = simpledialog.askstring("Input", "Ingrese el nombre del activo a registrar:").strip()
        if not nombre:
            messagebox.showerror("Error", "No se ingresó un nombre válido.")
            return

        filtro = df[df["Name"].str.lower() == nombre.lower()]

        if filtro.empty:
            messagebox.showerror("Error", f"No se encontró el activo con el nombre '{nombre}' en el archivo Excel.")
            return

        row = filtro.iloc[0].to_dict()
        session_token = self.obtener_token_sesion()

        location_id = self.obtener_location_id(session_token, row["Location"])
        manufacturer_id = self.obtener_manufacturer_id(session_token, row["Manufacturer"])

        if location_id is None or manufacturer_id is None:
            messagebox.showerror("Error", f"No se pudo encontrar la ubicación o el fabricante para el activo '{row['Name']}'")
            return

        # Preparar los datos para el registro en GLPI
        asset_data = {
            "name": row["Name"].strip(),
            "locations_id": location_id,
            "manufacturers_id": manufacturer_id,
            "serial": row["Serial Number"].strip(),
            #"otherserial": row["Inventory Number"].strip(),
            "comments": row["Comments"].strip(),
        }

        self.registrar_asset(session_token, asset_data, row["Asset Type"])

    def obtener_location_id(self, session_token, location_name):
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }
        
        params = {'searchText': location_name, 'range': '0-999'}
        response = requests.get(f"{GLPI_URL}/Location", headers=headers, params=params, verify=False)

        if response.status_code == 200:
            locations = response.json()
            for location in locations:
                if location.get("name", "").strip().lower() == location_name.strip().lower():
                    return location["id"]
        return None

    def obtener_manufacturer_id(self, session_token, manufacturer_name):
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }
        
        params = {'searchText': manufacturer_name, 'range': '0-999'}
        response = requests.get(f"{GLPI_URL}/Manufacturer", headers=headers, params=params, verify=False)

        if response.status_code == 200:
            manufacturers = response.json()
            for manufacturer in manufacturers:
                if manufacturer.get("name", "").strip().lower() == manufacturer_name.strip().lower():
                    return manufacturer["id"]
        return None

    def parse_qr_data_template(self, qr_string):
        asset_data = {}
        for line in qr_string.split("\n"):
            key, value = line.split(": ", 1)
            asset_data[key.strip()] = value.strip().replace('"', '')
        return asset_data

    def escanear_qr_con_celular(self):
        def run_capture(result_queue):
            ip_cam_url = IP_CAM_URL  # Cambiar por la URL de la cámara IP
            camera_open = True

            try:
                cap = cv2.VideoCapture(ip_cam_url)
                if not cap.isOpened():
                    result_queue.put(("error", "No se pudo acceder a la cámara del celular. Reintentalo..."))
                    cap.release()
                    cv2.destroyAllWindows()
                    return

                result_queue.put(("info", "Usando la cámara del celular. Presiona 'q' para salir."))

                while camera_open:
                    ret, frame = cap.read()
                    if not ret:
                        result_queue.put(("error", "Error al obtener el cuadro de la cámara. Reintentando conexión..."))
                        break  # Sale del bucle interno para reintentar la conexión

                    gray_frame = cv2.cvtColor(frame, cv2.COLOR_BGR2GRAY)
                    qr_codes = decode(gray_frame)

                    for qr in qr_codes:
                        qr_data = qr.data.decode('utf-8')
                        flag = self.es_codigo_valido(qr_data)
                        if flag in ["dell", "mac", "monitor"]:
                            result_queue.put(("info", f"Código QR {flag} escaneado: \n{qr_data}"))
                            result_queue.put(("data", qr_data))
                            camera_open = False
                            cap.release()
                            cv2.destroyAllWindows()
                            return
                        elif flag == "invalido":
                            # No mostrar mensaje repetitivo, solo continuar
                            continue

                    cv2.imshow("Escaneando QR con celular", frame)

                    if cv2.waitKey(1) & 0xFF == ord('q'):
                        camera_open = False
                        cap.release()
                        cv2.destroyAllWindows()
                        result_queue.put(("close", None))
                        return

                    # Verificar si la ventana fue cerrada
                    if cv2.getWindowProperty("Escaneando QR con celular", cv2.WND_PROP_VISIBLE) < 1:
                        camera_open = False
                        cap.release()
                        cv2.destroyAllWindows()
                        result_queue.put(("close", None))
                        return

                cap.release()
                cv2.destroyAllWindows()
            except Exception as e:
                result_queue.put(("error", f"Se produjo un error inesperado: {str(e)}"))
                cap.release()
                cv2.destroyAllWindows()

        result_queue = queue.Queue()
        capture_thread = threading.Thread(target=run_capture, args=(result_queue,))
        capture_thread.start()

        while True:
            try:
                msg_type, msg_content = result_queue.get(timeout=1)
                if msg_type == "info":
                    messagebox.showinfo("Información", msg_content)
                elif msg_type == "error":
                    messagebox.showerror("Error", msg_content)
                    return None
                elif msg_type == "data":
                    return msg_content
                elif msg_type == "close":
                    return None
            except queue.Empty:
                continue

    def es_codigo_valido(self, qr_data):
        # lógica para determinar si un código QR es válido
        # verificar si el código QR coincide con ciertos patrones
        patrones_validos = [
            # Dell Service Tag: 7 caracteres alfanuméricos (excluyendo I, O y Q)
            r'^(?!.*[IOQ])[A-HJ-NP-Z0-9]{7}$',  # Ejemplo: 8B9X1R3

            # Dell Express Service Code (conversión numérica del Service Tag)
            r'^\d{10}$',  # Ejemplo: 1234567890
            
            # MacBook Serial Number: Inicia con C02 o FVX seguido de 8-10 caracteres alfanuméricos
            r'^C02[A-Z0-9]{8,10}$',  # Ejemplo: C02X3Y5VFH5
            r'^FVX[A-Z0-9]{8,10}$',  # Ejemplo: FVXJ45KLD9

            # MacBook Serial Numbers con prefijo opcional 'S'
            r'^S?(C02|FVX)[A-Z0-9]{8,10}$',  # Soporta opcional 'S' delante del serial

            # MacBook Air/Pro: Formato reciente con 12 caracteres alfanuméricos
            r'^[A-Z0-9]{12}$',  # Ejemplo: W8P6W5T5YV3C

            # Dell Monitor Service Tag: "CN" seguido de 10 caracteres alfanuméricos
            r'^CN[A-Z0-9]{10}$',  # Ejemplo: CN0V7X9J129025AN2AX1

            # Dell Monitor etiqueta más corta con prefijo opcional "S"
            r'^S?[A-Z0-9]{7}$',  # Ejemplo: SCN12345

            # Monitores Dell con prefijo "SN" o "S/N"
            r'^(SN|S/N)\s*[A-Z0-9]{7,12}$',  # Ejemplo: SN 5JH34X1
        ]

        patron_valido_dell = [       
            # Dell Service Tag: 7 caracteres alfanuméricos (excluyendo I, O y Q)
            r'^(?!.*[IOQ])[A-HJ-NP-Z0-9]{7}$',  # Ejemplo: 8B9X1R3

            # Dell Express Service Code (conversión numérica del Service Tag)
            r'^\d{10}$',  # Ejemplo: 1234567890   
        ]
        
        patron_valido_mac = [
            # MacBook Serial Number: Inicia con C02 o FVX seguido de 8-10 caracteres alfanuméricos
            r'^C02[A-Z0-9]{8,10}$',  # Ejemplo: C02X3Y5VFH5
            r'^FVX[A-Z0-9]{8,10}$',  # Ejemplo: FVXJ45KLD9

            # MacBook Serial Numbers con prefijo opcional 'S'
            r'^S?(C02|FVX)[A-Z0-9]{8,10}$',  # Soporta opcional 'S' delante del serial

            # MacBook Air/Pro: Formato reciente con 12 caracteres alfanuméricos
            r'^[A-Z0-9]{10}$',  # Ejemplo: W8P6W5T5YV3
            r'^[A-Z0-9]{11}$',  # Ejemplo: W8P6W5T5YV3C
        ]
        
        patron_valido_monitor = [
            # Dell Monitor Service Tag: "CN" seguido de 10 caracteres alfanuméricos
            r'^CN[A-Z0-9]{10}$',  # Ejemplo: CN0V7X9J129025AN2AX1

            # Dell Monitor etiqueta más corta con prefijo opcional "S"
            r'^S?[A-Z0-9]{7}$',  # Ejemplo: SCN12345

            # Monitores Dell con prefijo "SN" o "S/N"
            r'^(SN|S/N)\s*[A-Z0-9]{7,12}$',  # Ejemplo: SN 5JH34X1
        ]

        for patron in patron_valido_dell:
            if re.match(patron, qr_data):
                return "dell"
        
        for patron in patron_valido_mac:
            if re.match(patron, qr_data):
                return "mac"
        
        for patron in patron_valido_monitor:
            if re.match(patron, qr_data):
                return "monitor"
        
        return "invalido"

    def limpiar_asset_data(self, asset_data):
        cleaned_data = {}
        for key, value in asset_data.items():
            if isinstance(value, float) and np.isnan(value):
                cleaned_data[key] = ""
            elif value is None:
                cleaned_data[key] = ""
            else:
                cleaned_data[key] = str(value).strip()
        return cleaned_data

    def procesar_archivo_excel(self, ruta_archivo):
        df = pd.read_excel(ruta_archivo, skiprows=0)
        df.columns = df.columns.str.strip()

        columnas_necesarias = [
            "Asset Type", "Name", "Location", "Manufacturer", "Model", "Serial Number", 
            "Inventory Number", "Comments", "Technician in Charge", 
            "Group in Charge", "Status", "Specific Fields (Dynamic Column)"
        ]

        for columna in columnas_necesarias:
            if columna not in df.columns:
                messagebox.showerror("Error", f"La columna '{columna}' no existe en el archivo Excel.")
                return

        df = df.fillna("").astype(str)

        session_token = self.obtener_token_sesion()
        if not session_token:
            messagebox.showerror("Error", "No se pudo obtener el token de sesión.")
            return

        # Mapeo para normalizar los tipos de assets
        asset_type_mapping = {
            "computer": "Computer",
            "network equipment": "Network Equipment",
            "consumables": "Consumables",
        }

        for index, row in df.iterrows():
            asset_type = row["Asset Type"].strip().lower()
            asset_type = asset_type_mapping.get(asset_type, None)

            if not asset_type:
                messagebox.showerror("Error", f"Tipo de asset desconocido: '{row['Asset Type']}' (Fila {index + 1}). Se omite esta fila.")
                continue  # Saltar la fila si el tipo de asset no es válido

            # Obtener location_id
            location_id = self.obtener_location_id(session_token, row["Location"].strip())
            if location_id is None:
                messagebox.showerror("Error", f"No se pudo encontrar la ubicación: {row['Location']} (Fila {index + 1})")
                continue  # Saltar la fila si no se encuentra la ubicación

            # Obtener manufacturer_id
            manufacturer_id = self.obtener_manufacturer_id(session_token, row["Manufacturer"].strip())
            #messagebox.showinfo("Información", f"ID del fabricante encontrado '{row['Manufacturer']}': {manufacturer_id}")
            if manufacturer_id is None:
                messagebox.showerror("Error", f"No se pudo encontrar el fabricante: {row['Manufacturer']} (Fila {index + 1})")
                continue  # Saltar la fila si no se encuentra el fabricante

            asset_data = {
                "name": row["Name"].strip(),
                "locations_id": location_id, 
                "manufacturers_id": manufacturer_id,
                "serial": row["Serial Number"].strip(),
                #"otherserial": row["Inventory Number"].strip(),
                "comments": row["Comments"].strip(),
            }

            messagebox.showinfo("Información", f"Procesando fila {index + 1}: {asset_data} como {asset_type}")
            self.registrar_asset(session_token, asset_data, asset_type)

    def salir(self):
        root.destroy()   

    # ---- Consumibles ------

    def agregar_consumible(self):
        
        messagebox.showinfo("Información", "--- Agregar Consumible al Stock ---")

        metodo = simpledialog.askstring("Input", "¿Desea escanear el QR o ingresar manualmente? (qr/manual): ").strip().lower()

        if metodo == "qr":
            qr_data = self.escanear_qr_con_celular()
            if qr_data:
                inventory_number = qr_data.strip()
                messagebox.showinfo("Información", f"Inventory Number detectado: {inventory_number}")
            else:
                messagebox.showerror("Error", "No se detectó ningún código QR.")
                return
        elif metodo == "manual":
            inventory_number = simpledialog.askstring("Input", "Ingrese el número de inventario o activo: ").strip()
        else:
            messagebox.showerror("Error", f"{metodo} no es un metodo valido.")
            return

        df = pd.read_excel(ruta_excel_consumibles)
        df.columns = df.columns.str.strip()  # Asegura que no haya espacios en los nombres de columnas

        # Verificar si el consumible ya está registrado en el Excel
        filtro = df[df["Inventory/Asset Number"].astype(str).str.lower() == inventory_number.lower()]

        if not filtro.empty:
            messagebox.showinfo("Información", f"Consumible con Inventory Number '{inventory_number}' encontrado en el Excel.")
            nombre_consumible = filtro.iloc[0]["Name"]
            location = filtro.iloc[0]["Location"]
        else:
            messagebox.showinfo("Información", f"No se encontró un consumible con el número de inventario '{inventory_number}'. Creando nuevo...")
            nombre_consumible = simpledialog.askstring("Input", "Ingrese el nombre del nuevo consumible: ").strip()
            location = simpledialog.askstring("Input", "Ingrese la ubicación del consumible: ").strip()

        cantidad = int(simpledialog.askstring("Input", "Ingrese la cantidad a agregar al stock: "))

        session_token = self.obtener_token_sesion()
        if not session_token:
            messagebox.showerror("Error", "No se pudo obtener el token de sesión.")
            return

        # Corrección: Se pasa tanto el nombre como el inventory_number a la función
        consumible_id = self.obtener_id_consumible(session_token, nombre_consumible, inventory_number)
        
        if not consumible_id:
            messagebox.showinfo("Información", f"No se encontró el consumible '{nombre_consumible}' en GLPI. Creando uno nuevo...")
            consumible_id = self.crear_consumible(session_token, nombre_consumible, inventory_number, location, cantidad)
            if not consumible_id:
                messagebox.showerror("Error", "Error al crear el consumible en GLPI.")
                return

        stock_actual = self.obtener_stock_actual(session_token, consumible_id)
        nuevo_stock = stock_actual + cantidad

        self.actualizar_stock_glpi(session_token, consumible_id, nuevo_stock)
        messagebox.showinfo("Información", f"Consumible '{nombre_consumible}' actualizado a {nuevo_stock} unidades en stock.")

        # Registrar en Excel
        self.actualizar_excel_consumible(nombre_consumible, inventory_number, location, nuevo_stock)

    def quitar_consumible(self):
        messagebox.showinfo("Información", "--- Quitar Consumible del Stock ---")

        metodo = simpledialog.askstring("Input", "¿Desea escanear el QR o ingresar manualmente? (qr/manual): ").strip().lower()

        if metodo == "qr":
            qr_data = self.escanear_qr_con_celular()
            if qr_data:
                inventory_number = qr_data.strip()
                messagebox.showinfo("Información", f"Inventory Number detectado: {inventory_number}")
            else:
                messagebox.showerror("Error", "No se detectó ningún código QR.")
                return
        else:
            inventory_number = simpledialog.askstring("Input", "Ingrese el número de inventario o activo: ").strip()

        df = pd.read_excel(ruta_excel_consumibles)
        df.columns = df.columns.str.strip()  # Asegurar que no haya espacios en los nombres de columnas

        # Buscar el consumible en el Excel por inventory_number
        filtro = df[df["Inventory/Asset Number"].astype(str).str.lower() == inventory_number.lower()]

        if filtro.empty:
            messagebox.showerror("Error", f"No se encontró el consumible con Inventory Number '{inventory_number}' en el archivo Excel.")
            return

        # Obtener datos existentes del consumible
        nombre_consumible = filtro.iloc[0]["Name"]
        location = filtro.iloc[0]["Location"]
        stock_actual = int(filtro.iloc[0]["Stock Target"])

        cantidad = int(simpledialog.askstring("Input", f"Ingrese la cantidad a retirar (Stock actual: {stock_actual}): "))

        if stock_actual < cantidad:
            messagebox.showerror("Error", "No se puede retirar más cantidad de la que hay en stock.")
            return

        nuevo_stock = stock_actual - cantidad

        session_token = self.obtener_token_sesion()
        if not session_token:
            messagebox.showerror("Error", "No se pudo obtener el token de sesión.")
            return

        consumible_id = self.obtener_id_consumible(session_token, nombre_consumible, inventory_number)
        if not consumible_id:
            messagebox.showerror("Error", f"No se encontró el consumible '{nombre_consumible}' en GLPI.")
            return

        self.actualizar_stock_glpi(session_token, consumible_id, nuevo_stock)
        messagebox.showinfo("Información", f"Consumible '{nombre_consumible}' actualizado a {nuevo_stock} unidades en stock.")

        # Actualizar en Excel
        self.actualizar_excel_consumible(nombre_consumible, inventory_number, location, nuevo_stock)

    def crear_consumible(self, session_token, nombre, inventory_number, location, stock_target):
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        payload = {
            "input": {
                "name": nombre,
                "otherserial": inventory_number,
                "locations_id": self.obtener_location_id(session_token, location),
                "stock_target": stock_target
            }
        }

        response = requests.post(f"{GLPI_URL}/ConsumableItem", headers=headers, json=payload, verify=False)

        if response.status_code == 201:
            consumible_id = response.json().get("id")
            messagebox.showinfo("Información", f"Consumible '{nombre}' creado exitosamente con ID {consumible_id}.")
            return consumible_id
        else:
            messagebox.showerror("Error", f"Error al crear el consumible: {response.status_code}")
            try:
                messagebox.showerror("Error", response.json())
            except json.JSONDecodeError:
                messagebox.showerror("Error", response.text)
            return None

    def obtener_id_consumible(self, session_token, nombre_consumible, inventory_number):
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        params = {
            "searchText": nombre_consumible.strip().lower(),
            "range": "0-999"
        }

        response = requests.get(f"{GLPI_URL}/ConsumableItem", headers=headers, params=params, verify=False)
        
        if response.status_code == 200:
            consumibles = response.json()
            for consumible in consumibles:
                # Convertir a cadena de texto y limpiar espacios en blanco
                consumible_name = str(consumible.get("name", "")).strip().lower()
                consumible_serial = str(consumible.get("otherserial", "")).strip().lower()
                inventory_number_str = str(inventory_number).strip().lower()

                if consumible_name == nombre_consumible.strip().lower() and consumible_serial == inventory_number_str:
                    return consumible["id"]

        messagebox.showinfo("Información", f"No se encontró el consumible con nombre '{nombre_consumible}' y número de inventario '{inventory_number}' en GLPI.")
        return None

    def obtener_stock_actual(self, session_token, consumible_id):
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        response = requests.get(f"{GLPI_URL}/ConsumableItem/{consumible_id}", headers=headers, verify=False)

        if response.status_code == 200:
            return int(response.json().get("stock_target", 0))
        else:
            messagebox.showerror("Error", f"Error al obtener el stock del consumible ID {consumible_id}: {response.status_code}")
            return 0

    def actualizar_stock_glpi(self, session_token, consumible_id, nuevo_stock):
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        payload = {
            "input": {
                "id": consumible_id,
                "stock_target": nuevo_stock
            }
        }

        response = requests.put(f"{GLPI_URL}/ConsumableItem/{consumible_id}", headers=headers, json=payload, verify=False)

        if response.status_code == 200:
            messagebox.showinfo("Información", f"Stock del consumible ID {consumible_id} actualizado correctamente a {nuevo_stock}.")
        else:
            messagebox.showerror("Error", f"Error al actualizar el stock en GLPI: {response.status_code}")
            try:
                messagebox.showerror("Error", response.json())
            except json.JSONDecodeError:
                messagebox.showerror("Error", response.text)

    def actualizar_excel_consumible(self, nombre, inventory_number, location, stock):
        df = pd.read_excel(ruta_excel_consumibles)
        df.columns = df.columns.str.strip()  # Asegura nombres sin espacios extra

        # Convertir inventory_number a string para evitar errores
        inventory_number_str = str(inventory_number).strip().lower()

        # Verificar si el consumible ya está registrado
        filtro = df[
            (df["Name"].str.lower() == nombre.lower()) & 
            (df["Inventory/Asset Number"].astype(str).str.lower() == inventory_number_str)
        ]

        if not filtro.empty:
            df.loc[
                (df["Name"].str.lower() == nombre.lower()) &
                (df["Inventory/Asset Number"].astype(str).str.lower() == inventory_number_str), 
                "Stock Target"
            ] = stock
        else:
            nuevo_consumible = pd.DataFrame([{
                "Name": nombre,
                "Inventory/Asset Number": inventory_number_str,  # Convertir a string
                "Location": location,
                "Stock Target": stock
            }])
            df = pd.concat([df, nuevo_consumible], ignore_index=True)

        df.to_excel(ruta_excel_consumibles, index=False)
        messagebox.showinfo("Información", f"El consumible '{nombre}' ha sido registrado/actualizado en el Excel.")

    # ------ Monitor --------

    def manejar_qr_monitor(self):
        qr_data = self.escanear_qr_con_celular()
        if qr_data:
            if re.match(r'^CN[A-Z0-9]{10}$', qr_data) or re.match(r'^S?[A-Z0-9]{7}$', qr_data) or re.match(r'^(SN|S/N)\s*[A-Z0-9]{7,12}$', qr_data):
                messagebox.showinfo("Información", "Monitor detectado. Procesando datos...")

                serial_number = qr_data.strip()

                if serial_number:
                    messagebox.showinfo("Información", f"Serial Number detectado: {serial_number}")
                    confirmacion = simpledialog.askstring("Confirmación", "¿Es correcto este Serial Number, desea continuar? (sí/no): ").strip().lower()
                    if confirmacion not in ["sí", "si", "Si", "Sí"]:
                        messagebox.showinfo("Información", "Operación cancelada por el usuario.")
                        return
                    else:
                        if self.verificar_existencia_en_excel(serial_number):
                            messagebox.showinfo("Información", f"El activo con serial '{serial_number}' ya está registrado en el Excel. No se agregará.")
                            return
                        asset_data = self.procesar_qr_monitor(serial_number)
                        self.agregar_a_excel(asset_data)
            else:
                messagebox.showerror("Error", "Código QR no corresponde a un monitor.")
        else:
            messagebox.showerror("Error", "No se detectó ningún código QR.")

    def procesar_qr_monitor(self, qr_data):
        # Plantilla para monitores
        plantilla_monitor = {
            "Asset Type": "Monitor",
            "Status": "Stocked",  # Definir un estado predeterminado
            "User": None,  # Pedir al usuario
            "Name": None,  # Generado automáticamente
            "Location": None,  # Pedir al usuario
            "Manufacturer": "Dell Inc.",  # Extraído del QR
            "Model": None,  # Extraído del QR
            "Serial Number": qr_data.strip(),  # Código QR escaneado
            "Comments": "Check",
        }

        # Solicitar datos adicionales al usuario
        plantilla_monitor["Location"] = simpledialog.askstring("Input", "Ingrese la ubicación del monitor:").strip()
        #plantilla_monitor["Manufacturer"] = simpledialog.askstring("Input", "Ingrese el fabricante del monitor:").strip()
        #plantilla_monitor["Model"] = simpledialog.askstring("Input", "Ingrese el modelo del monitor:").strip()

        # Generar el nombre del activo a partir del modelo
        plantilla_monitor["Name"] = f"{plantilla_monitor['Model']}-{plantilla_monitor['Serial Number']}"

        return plantilla_monitor

    def entregar_monitor(self):
        messagebox.showinfo("Información", "--- Entregar Monitor a Usuario ---")
        metodo = simpledialog.askstring("Input", "¿Desea escanear el QR o ingresar el número de serie manualmente? (escanear/manual):").strip().lower()

        if metodo == "escanear":
            qr_data = self.escanear_qr_con_celular()
            if re.match(r'^CN[A-Z0-9]{10}$', qr_data) or re.match(r'^S?[A-Z0-9]{7}$', qr_data) or re.match(r'^(SN|S/N)\s*[A-Z0-9]{7,12}$', qr_data):
                messagebox.showinfo("Información", "Monitor detectado. Procesando datos...")
                serial_number = qr_data.strip()
            else:
                messagebox.showerror("Error", "Código QR no corresponde a un monitor válido.")
                return
        elif metodo == "manual":
            res = simpledialog.askstring("Input", "Ingrese el número de serie del monitor:").strip()
            if re.match(r'^CN[A-Z0-9]{10}$', res) or re.match(r'^S?[A-Z0-9]{7}$', res) or re.match(r'^(SN|S/N)\s*[A-Z0-9]{7,12}$', res):
                serial_number = res
                messagebox.showinfo("Información", "Monitor detectado. Procesando datos...")
            else:
                messagebox.showerror("Error", "Código QR no corresponde a un monitor válido.")
                return
                
        else:
            messagebox.showerror("Error", "Método no válido. Intente nuevamente.")
            return

        df = pd.read_excel(ruta_excel)
        if df.empty:
            messagebox.showerror("Error", "El archivo Excel está vacío.")
            return

        filtro = df[df["Serial Number"].str.lower() == serial_number.lower()]

        if filtro.empty:
            messagebox.showerror("Error", f"No se encontró un monitor con el número de serie '{serial_number}' en el archivo Excel.")
            return

        nuevo_usuario = simpledialog.askstring("Input", "Ingrese el nombre del usuario que recibirá el monitor:").strip()

        # Manejar valores NaN antes de actualizar el DataFrame
        df["User"] = df["User"].fillna("")
        df["Name"] = df["Name"].fillna("Unknown")

        # Determinar el nuevo nombre del monitor en base al usuario
        manufacturer = filtro["Manufacturer"].values[0]
        if manufacturer == "Dell" or manufacturer == "dell" or manufacturer == "Dell inc." or manufacturer == "Dell Inc." or manufacturer == "DELL":
            new_name = f"{nuevo_usuario}-DellMonitor"
        elif manufacturer == "Samsung":
            new_name = f"{nuevo_usuario}-SamsungMonitor"
        else:
            new_name = f"{nuevo_usuario}-Monitor"

        df.loc[df["Serial Number"].str.lower() == serial_number.lower(), "User"] = nuevo_usuario
        df.loc[df["Serial Number"].str.lower() == serial_number.lower(), "Name"] = new_name

        df.to_excel(ruta_excel, index=False)
        messagebox.showinfo("Información", f"Monitor con número de serie '{serial_number}' asignado a '{nuevo_usuario}' en el Excel.")

        # Actualizar en GLPI
        session_token = self.obtener_token_sesion()
        if not session_token:
            messagebox.showerror("Error", "No se pudo obtener el token de sesión.")
            return

        asset_id = self.obtener_asset_id_por_serial_monitor(session_token, serial_number)
        if not asset_id:
            messagebox.showerror("Error", "No se pudo encontrar el activo en GLPI.")
            return

        asset_data = filtro.iloc[0].to_dict()
        asset_data["User"] = nuevo_usuario
        asset_data["Name"] = new_name  

        self.actualizar_asset_glpi_monitor(session_token, asset_id, asset_data)

    def actualizar_asset_glpi_monitor(self, session_token, asset_id, asset_data):
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        # Obtener el ID del usuario basado en su nombre
        user_id = self.obtener_user_id(session_token, asset_data["User"])
        messagebox.showinfo("Información", f"ID del usuario encontrado: {user_id}")
        if not user_id:
            messagebox.showerror("Error", f"No se encontró el usuario '{asset_data['User']}' en GLPI.")
            return

        # Determinar el nuevo nombre del monitor
        if "Dell" in asset_data["Manufacturer"]:
            new_name = f"{asset_data['User']}-DellMonitor"
        elif "Samsung" in asset_data["Manufacturer"]:
            new_name = f"{asset_data['User']}-SamsungMonitor"
        else:
            new_name = f"{asset_data['User']}-Monitor"

        # Preparar datos para la actualización en GLPI
        payload = {
            "input": {
                "id": asset_id,  
                "name": new_name,
                "users_id": user_id
            }
        }

        response = requests.put(f"{GLPI_URL}/Monitor/{asset_id}", headers=headers, json=payload, verify=False)

        if response.status_code == 200:
            messagebox.showinfo("Éxito", f"Monitor con ID {asset_id} actualizado correctamente en GLPI con el nombre '{new_name}'.")
        else:
            messagebox.showerror("Error", f"Error al actualizar el monitor en GLPI: {response.status_code}")
            try:
                messagebox.showerror("Error", response.json())
            except json.JSONDecodeError:
                messagebox.showerror("Error", response.text)

    def obtener_id_por_nombre_monitor(self, session_token, asset_name):
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        params = {
            "searchText": asset_name.strip().lower(),
            "range": "0-999"
        }

        response = requests.get(f"{GLPI_URL}/Monitor", headers=headers, params=params, verify=False)

        if response.status_code == 200:
            for asset in response.json():
                if asset.get("name").strip().lower() == asset_name.strip().lower():
                    messagebox.showinfo("Información", f"Monitor encontrado: {asset_name}, ID: {asset.get('id')}")
                    return asset.get("id")

            messagebox.showinfo("Información", f"No se encontró el ID para el monitor '{asset_name}' en GLPI.")
            return None
        else:
            messagebox.showerror("Error", f"Error al buscar el ID del monitor: {response.status_code}")
            return None

    def obtener_asset_id_por_serial_monitor(self, session_token, serial_number):
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        params = {
            "searchText": serial_number.strip().lower(),
            "range": "0-999"
        }

        response = requests.get(f"{GLPI_URL}/search/Monitor", headers=headers, params=params, verify=False)

        if response.status_code == 200:
            assets = response.json().get("data", [])
            
            for asset in assets:
                serial_found = (asset.get("5") or "").strip().lower()  # Clave 5 es el serial number
                asset_name = asset.get("1")  # Clave 1 es el nombre del asset
                
                if serial_found == serial_number.lower():
                    messagebox.showinfo("Información", f"Monitor encontrado: {asset_name}, Serial: {serial_number}")
                    return self.obtener_id_por_nombre_monitor(session_token, asset_name)

            messagebox.showinfo("Información", f"No se encontró un monitor con el número de serie '{serial_number}' en GLPI.")
            return None
        else:
            messagebox.showerror("Error", f"Error al buscar el monitor en GLPI: {response.status_code}")
            return None

    ## ----- Laptops -------

    def procesar_qr_laptop(self, flag, qr_data):
        if flag == "Dell":
            # Plantilla de la laptop Dell Latitude
            plantilla_dell = {
                "Asset Type": "Computer",
                "Status": "Stocked",  # Solicitar por pantalla
                "User": None,    # Solicitar por pantalla
                "Name": None,    # Generado a partir del nombre del usuario
                "Computer Types": "Laptop",
                "Location": None,  # Solicitar por pantalla
                "Manufacturer": "Dell inc.",
                "Model": "Latitude",
                "Serial Number": qr_data.strip(),  # QR escaneado de la laptop
                "Comments": "Check",
            }

            # Solicitar datos adicionales al usuario
            #plantilla_dell["Status"] = simpledialog.askstring("Input", "Ingrese el estado del activo (Activo/Inactivo):").strip()
            #plantilla_dell["User"] = simpledialog.askstring("Input", "Ingrese el nombre del usuario:").strip()
            plantilla_dell["Location"] = simpledialog.askstring("Input", "Ingrese la ubicación del activo:").strip()
            
            # Generar el nombre del activo a partir del usuario
            plantilla_dell["Name"] = f"{plantilla_dell['User']}-Latitude"

            return plantilla_dell
        elif flag == "Mac":
            # Plantilla para laptops MacBook
            plantilla_mac = {
                "Asset Type": "Computer",
                "Status": "Stocked",  # Solicitar por pantalla
                "User": None,    # Solicitar por pantalla
                "Name": None,    # Generado a partir del nombre del usuario
                "Computer Types": "Laptop",
                "Location": None,  # Solicitar por pantalla
                "Manufacturer": "Apple Inc",
                "Model": "MacBook Pro",
                "Serial Number": qr_data.strip(),  # QR escaneado de la laptop
                "Comments": "Check",
            }

            # Solicitar datos adicionales al usuario
            #plantilla_mac["Status"] = simpledialog.askstring("Input", "Ingrese el estado del activo (Activo/Inactivo):").strip()
            #plantilla_mac["User"] = simpledialog.askstring("Input", "Ingrese el nombre del usuario:").strip()
            plantilla_mac["Location"] = simpledialog.askstring("Input", "Ingrese la ubicación del activo:").strip()
            
            # Generar el nombre del activo a partir del usuario
            plantilla_mac["Name"] = f"{plantilla_mac['User']}-MacBookPro"

            return plantilla_mac

    def guardar_plantilla_txt(self, asset_data, nombre_archivo):
        with open(nombre_archivo, 'w') as file:
            for key, value in asset_data.items():
                file.write(f"{key}: {value}\n")
        messagebox.showinfo("Información", f"Plantilla guardada en {nombre_archivo}")

    def manejar_qr_laptop(self, flag):
        try: 
            if flag == "Register":
                manufacturer = simpledialog.askstring("Input", "Ingrese el fabricante del laptop (Dell/Mac):").strip().lower()
                serial_number = None
                
                if manufacturer == "Dell" or manufacturer == "dell" or manufacturer == "Dell inc." or manufacturer == "Dell Inc." or manufacturer == "DELL":
                    metodo = simpledialog.askstring("Input", "¿Desea escanear el QR o ingresar el Service Tag manualmente? (escanear/manual):").strip().lower()
                    if metodo == "escanear":
                        qr_data = self.escanear_qr_con_celular()
                        if re.match(r'\bcs[a-z0-9]{5}\b', qr_data) or re.match(r'^[A-Za-z0-9]{7}$', qr_data):
                            messagebox.showinfo("Información", "Laptop Dell detectada. Procesando datos...")
                            serial_number = qr_data
                            confirmacion = simpledialog.askstring("Confirmación", f"¿Es correcto este Serial Number: {serial_number} desea continuar? (sí/no):").strip().lower()
                            if confirmacion in ["sí", "si", "Si", "Sí"]:
                                if self.verificar_existencia_en_excel(serial_number):
                                    messagebox.showinfo("Información", f"El activo con serial '{serial_number}' ya está registrado en el Excel. No se agregará.")
                                    return
                                else:
                                    asset_data = self.procesar_qr_laptop("Dell", serial_number)
                                    self.agregar_a_excel(asset_data)
                            else: 
                                messagebox.showinfo("Información", "Operación cancelada por el usuario.")
                                return
                                
                    elif metodo == "manual":
                        serial_number = simpledialog.askstring("Input", "Ingrese el Service Tag del laptop:").strip()
                        if re.match(r'^[A-Za-z0-9]{7}$', serial_number) or re.match(r'\bcs[a-z0-9]{5}\b', serial_number):
                            messagebox.showinfo("Información", "Laptop Dell detectada. Procesando datos...")
                            confirmacion = simpledialog.askstring("Confirmación", f"¿Es correcto este Serial Number: {serial_number} desea continuar? (sí/no):").strip().lower()
                            if confirmacion in ["sí", "si", "Si", "Sí"]:
                                if self.verificar_existencia_en_excel(serial_number):
                                    messagebox.showinfo("Información", f"El activo con serial '{serial_number}' ya está registrado en el Excel. No se agregará.")
                                    return
                                else:
                                    asset_data = self.procesar_qr_laptop("Dell", serial_number)
                                    self.agregar_a_excel(asset_data)
                            else: 
                                messagebox.showinfo("Información", "Operación cancelada por el usuario.")
                                return
                        else:
                            messagebox.showerror("Error", "Service Tag no válido. Intente nuevamente.")
                            return
                    
                    else: 
                        messagebox.showerror("Error", "Método no válido. Intente nuevamente.")
                        return 
                elif manufacturer == "Mac" or manufacturer == "mac" or manufacturer == "Mac Inc." or manufacturer == "Apple Inc." or manufacturer == "APPLE":
                    metodo = simpledialog.askstring("Input", "¿Desea escanear el QR o ingresar el Serial Number manualmente? (escanear/manual):").strip().lower()
                    if metodo == "escanear":
                        qr_data = self.escanear_qr_con_celular()
                        if re.match(r'\bC02[A-Za-z0-9]{8,10}\b', qr_data) or re.match(r'^[A-Za-z0-9]{10,12}$', qr_data) or re.match(r'^S?[C02][A-Za-z0-9]{8}$', qr_data) or re.match(r'^S[A-Za-z0-9]{9}$', qr_data):
                            messagebox.showinfo("Información", "Laptop MacBook detectada. Procesando datos...")
                            # Remover la 'S' del serial number si existe al inicio
                            serial_number = qr_data
                            serial_number = serial_number[1:] if serial_number.startswith("S") else serial_number
                            confirmacion = simpledialog.askstring("Confirmación", f"¿Es correcto este Serial Number: {serial_number} desea continuar? (sí/no):").strip().lower()
                            if confirmacion in ["sí", "si", "Si", "Sí"]:
                                if self.verificar_existencia_en_excel(serial_number):
                                    messagebox.showinfo("Información", f"El activo con serial '{serial_number}' ya está registrado en el Excel. No se agregará.")
                                    return
                                else:
                                    asset_data = self.procesar_qr_laptop("Mac", serial_number)
                                    self.agregar_a_excel(asset_data)
                            else: 
                                messagebox.showinfo("Información", "Operación cancelada por el usuario.")
                                return
                    elif metodo == "manual": 
                        serial_number = simpledialog.askstring("Input", "Ingrese el Service Tag del laptop:").strip()
                        if re.match(r'\bC02[A-Za-z0-9]{8,10}\b', serial_number) or re.match(r'^[A-Za-z0-9]{10,12}$', serial_number) or re.match(r'^S?[C02][A-Za-z0-9]{8}$', serial_number) or re.match(r'^S[A-Za-z0-9]{9}$', serial_number):
                            messagebox.showinfo("Información", "Laptop Mac detectada. Procesando datos...")
                            confirmacion = simpledialog.askstring("Confirmación", f"¿Es correcto este Serial Number: {serial_number} desea continuar? (sí/no):").strip().lower()
                            if confirmacion in ["sí", "si", "Si", "Sí"]:
                                if self.verificar_existencia_en_excel(serial_number):
                                    messagebox.showinfo("Información", f"El activo con serial '{serial_number}' ya está registrado en el Excel. No se agregará.")
                                    return
                                else:
                                    asset_data = self.procesar_qr_laptop("Mac", serial_number)
                                    self.agregar_a_excel(asset_data)
                            else:
                                messagebox.showinfo("Información", "Operación cancelada por el usuario.")
                                return
                        else:
                            messagebox.showerror("Error", "Service Tag no válido. Intente nuevamente.")
                            return
                else: 
                    messagebox.showerror("Error", "Fabricante no válido. Intente nuevamente.")
                    return
            elif flag == "Deliver":
                manufacturer = simpledialog.askstring("Input", "Ingrese el fabricante del laptop (Dell/Mac):").strip().lower()
                serial_number = None
                
                if manufacturer == "Dell" or manufacturer == "dell" or manufacturer == "Dell inc." or manufacturer == "Dell Inc." or manufacturer == "DELL":
                    metodo = simpledialog.askstring("Input", "¿Desea escanear el QR o ingresar el Service Tag manualmente? (escanear/manual):").strip().lower()
                    if metodo == "escanear":
                        qr_data = self.escanear_qr_con_celular()
                        if re.match(r'\bcs[a-z0-9]{5}\b', qr_data) or re.match(r'^[A-Za-z0-9]{7}$', qr_data):
                            messagebox.showinfo("Información", "Laptop Dell detectada. Procesando datos...")
                            serial_number = qr_data
                            confirmacion = simpledialog.askstring("Confirmación", f"¿Es correcto este Serial Number: {serial_number} desea continuar? (sí/no):").strip().lower()
                            if confirmacion in ["sí", "si", "Si", "Sí"]:
                                return serial_number, manufacturer
                            else:
                                messagebox.showinfo("Información", "Operación cancelada por el usuario.")
                                return
                    elif metodo == "manual":
                        serial_number = simpledialog.askstring("Input", "Ingrese el Service Tag del laptop:").strip()
                        if re.match(r'^[A-Za-z0-9]{7}$', serial_number) or re.match(r'\bcs[a-z0-9]{5}\b', serial_number):
                            messagebox.showinfo("Información", "Laptop Dell detectada. Procesando datos...")
                            confirmacion = simpledialog.askstring("Confirmación", f"¿Es correcto este Serial Number: {serial_number} desea continuar? (sí/no):").strip().lower()
                            if confirmacion in ["sí", "si", "Si", "Sí"]:
                                return serial_number, manufacturer
                            else: 
                                messagebox.showinfo("Información", "Operación cancelada por el usuario.")
                                return
                        else:
                            messagebox.showerror("Error", "Service Tag no válido. Intente nuevamente.")
                            return
                    else: 
                        messagebox.showerror("Error", "Método no válido. Intente nuevamente.")
                        return 
                    
                elif manufacturer == "Mac" or manufacturer == "mac" or manufacturer == "Mac Inc." or manufacturer == "Apple Inc." or manufacturer == "APPLE":
                    metodo = simpledialog.askstring("Input", "¿Desea escanear el QR o ingresar el Serial Number manualmente? (escanear/manual):").strip().lower()
                    if metodo == "escanear":
                        qr_data = self.escanear_qr_con_celular()
                        if re.match(r'\bC02[A-Za-z0-9]{8,10}\b', qr_data) or re.match(r'^[A-Za-z0-9]{10,12}$', qr_data) or re.match(r'^S?[C02][A-Za-z0-9]{8}$', qr_data) or re.match(r'^S[A-Za-z0-9]{9}$', qr_data):
                            messagebox.showinfo("Información", "Laptop MacBook detectada. Procesando datos...")
                            # Remover la 'S' del serial number si existe al inicio
                            serial_number = qr_data
                            serial_number = serial_number[1:] if serial_number.startswith("S") else serial_number
                            confirmacion = simpledialog.askstring("Confirmación", f"¿Es correcto este Serial Number: {serial_number} desea continuar? (sí/no):").strip().lower()
                            if confirmacion in ["sí", "si", "Si", "Sí"]:
                                return serial_number, manufacturer
                            else:
                                messagebox.showinfo("Información", "Operación cancelada por el usuario.")
                                return
                    elif metodo == "manual": 
                        serial_number = simpledialog.askstring("Input", "Ingrese el Service Tag del laptop:").strip()
                        if re.match(r'\bC02[A-Za-z0-9]{8,10}\b', serial_number) or re.match(r'^[A-Za-z0-9]{10,12}$', serial_number) or re.match(r'^S?[C02][A-Za-z0-9]{8}$', serial_number) or re.match(r'^S[A-Za-z0-9]{9}$', serial_number):
                            messagebox.showinfo("Información", "Laptop Mac detectada. Procesando datos...")
                            confirmacion = simpledialog.askstring("Confirmación", f"¿Es correcto este Serial Number: {serial_number} desea continuar? (sí/no):").strip().lower()
                            if confirmacion in ["sí", "si", "Si", "Sí"]:
                                return serial_number, manufacturer 
                            else:
                                messagebox.showinfo("Información", "Operación cancelada por el usuario.")
                                return    
                        else:
                            messagebox.showerror("Error", "Service Tag no válido. Intente nuevamente.")
                            return
                    else: 
                        messagebox.showerror("Error", "Método no válido. Intente nuevamente.")
                        return 
                        
        except Exception as e:
            messagebox.showerror("Error", f"Se produjo un error inesperado: {str(e)}")
            return
            
    def registrar_laptop(self):
        try:
            messagebox.showinfo("Información", "--- Registrar Laptop en Excel ---")
            self.manejar_qr_laptop("Register")
        except Exception as e:
            messagebox.showerror("Error", f"Se produjo un error inesperado: {str(e)}")
            
    def entregar_laptop(self):
        try: 
            messagebox.showinfo("Información", "--- Entregar Laptop a Usuario ---")
            serial_number, manufacturer = self.manejar_qr_laptop("Deliver")

            # Cargar el archivo Excel
            df = pd.read_excel(ruta_excel)
            if df.empty:
                messagebox.showerror("Error", "El archivo Excel está vacío.")
                return

            # Validar columnas necesarias
            required_columns = ["Serial Number", "Manufacturer", "User", "Name"]
            if not all(col in df.columns for col in required_columns):
                messagebox.showerror("Error", "El archivo Excel no contiene las columnas necesarias.")
                return
            
            filtro = df[df["Serial Number"].str.lower() == serial_number.lower()]

            if filtro.empty:
                messagebox.showerror("Error", f"No se encontró un laptop con el Service Tag '{serial_number}' en el archivo Excel.")
                return

            nuevo_usuario = simpledialog.askstring("Input", "Ingrese el nombre del usuario que recibirá el laptop:").strip()
            if not nuevo_usuario:
                messagebox.showerror("Error", "El nombre del usuario no puede estar vacío.")
                return
            
            # Manejar valores NaN antes de actualizar el DataFrame
            df["User"] = df["User"].fillna("")
            df["Name"] = df["Name"].fillna("Unknown")


            # Determinar el nuevo nombre del laptop en base al fabricante
            if manufacturer == "Dell" or manufacturer == "dell" or manufacturer == "Dell inc." or manufacturer == "Dell Inc." or manufacturer == "DELL":
                new_name = f"{nuevo_usuario}-Latitude"
            elif manufacturer == "Mac" or manufacturer == "mac" or manufacturer == "Mac Inc." or manufacturer == "Apple Inc." or manufacturer == "Apple":
                new_name = f"{nuevo_usuario}-MacBookPro"
            else:
                messagebox.showerror("Error", "No se pudo determinar el fabricante del laptop.")
                return

            # Actualizar DataFrame con los nuevos valores
            df.loc[df["Serial Number"].str.lower() == serial_number.lower(), "User"] = nuevo_usuario
            df.loc[df["Serial Number"].str.lower() == serial_number.lower(), "Name"] = new_name

            df.to_excel(ruta_excel, index=False)
            messagebox.showinfo("Información", f"Laptop con Service Tag '{serial_number}' asignado a '{nuevo_usuario}' en el Excel.")

            # Actualizar en GLPI
            session_token = self.obtener_token_sesion()
            if not session_token:
                messagebox.showerror("Error", "No se pudo obtener el token de sesión.")
                return

            asset_id = self.obtener_asset_id_por_serial(session_token, serial_number)
            if not asset_id:
                messagebox.showerror("Error", "No se pudo encontrar el activo en GLPI.")
                return

            asset_data = filtro.iloc[0].to_dict()
            asset_data["User"] = nuevo_usuario
            asset_data["Name"] = new_name  

            self.actualizar_asset_glpi(session_token, asset_id, asset_data)
        except Exception as e:
            messagebox.showerror("Error", f"Se produjo un error inesperado: {str(e)}")

    def obtener_asset_id_por_serial(self, session_token, serial_number):
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        params = {
            "searchText": serial_number.strip().lower(),
            "range": "0-999"
        }

        # Buscar el asset por número de serie
        response = requests.get(f"{GLPI_URL}/search/Computer", headers=headers, params=params, verify=False)

        if response.status_code == 200:
            assets = response.json().get("data", [])
            
            for asset in assets:
                serial_found = (asset.get("5") or "").strip().lower()  # Clave 5 es el serial number
                asset_name = asset.get("1")  # Clave 1 es el nombre del asset
                
                if serial_found == serial_number.lower():
                    messagebox.showinfo("Información", f"Activo encontrado: {asset_name}, Serial: {serial_number}")
                    # Ahora buscamos el ID utilizando el nombre del activo encontrado
                    return self.obtener_id_por_nombre(session_token, asset_name)

            messagebox.showinfo("Información", f"No se encontró un activo con el serial number '{serial_number}' en GLPI.")
            return None
        else:
            messagebox.showerror("Error", f"Error al buscar el activo en GLPI: {response.status_code}")
            return None

    def obtener_id_por_nombre(self, session_token, asset_name):
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        params = {
            "searchText": asset_name.strip().lower(),
            "range": "0-999"
        }

        # Buscar el asset por nombre
        response = requests.get(f"{GLPI_URL}/Computer", headers=headers, params=params, verify=False)

        if response.status_code == 200:
            for asset in response.json():
                if asset.get("name").strip().lower() == asset_name.strip().lower():
                    messagebox.showinfo("Información", f"Activo encontrado: {asset_name}, ID: {asset.get('id')}")
                    return asset.get("id")

            messagebox.showinfo("Información", f"No se encontró el ID para el activo '{asset_name}' en GLPI.")
            return None
        else:
            messagebox.showerror("Error", f"Error al buscar el ID del activo: {response.status_code}")
            return None

    def actualizar_asset_glpi(self, session_token, asset_id, asset_data):
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        # Obtener el ID del usuario basado en su nombre
        user_id = self.obtener_user_id(session_token, asset_data["User"])
        messagebox.showinfo("Información", f"ID del usuario encontrado: {user_id}")
        if not user_id:
            messagebox.showerror("Error", f"No se encontró el usuario '{asset_data['User']}' en GLPI.")
            return

        # Determinar el nuevo nombre según el fabricante
        if "Dell" in asset_data["Manufacturer"] or "Dell Inc." in asset_data["Manufacturer"] or "dell" in asset_data["Manufacturer"] or "DELL" in asset_data["Manufacturer"] or "Dell inc." in asset_data["Manufacturer"]:
            new_name = f"{asset_data['User']}-Latitude"
        elif "Apple" in asset_data["Manufacturer"] or "Apple Inc." in asset_data["Manufacturer"] or "apple" in asset_data["Manufacturer"] or "MAC" in asset_data["Manufacturer"] or "Mac" in asset_data["Manufacturer"] or "mac" in asset_data["Manufacturer"]:
            new_name = f"{asset_data['User']}-MacBookPro"
        else:
            messagebox.showerror("Error", "No se pudo determinar el fabricante del laptop.")
            return

        # Preparar datos para la actualización en GLPI
        payload = {
            "input": {
                "id": asset_id,  
                "name": new_name,
                #"users_id": user_id
            }
        }

        response = requests.put(f"{GLPI_URL}/Computer/{asset_id}", headers=headers, json=payload, verify=False)

        if response.status_code == 200:
            messagebox.showinfo("Éxito", f"Activo con ID {asset_id} actualizado correctamente en GLPI con el nombre '{new_name}'.")
        else:
            messagebox.showerror("Error", f"Error al actualizar el activo en GLPI: {response.status_code}")
            try:
                messagebox.showerror("Error", response.json())
            except json.JSONDecodeError:
                messagebox.showerror("Error", response.text)

    def obtener_user_id(self, session_token, username):
        headers = {
            "Content-Type": "application/json",
            "Session-Token": session_token,
            "App-Token": APP_TOKEN
        }

        params = {"searchText": username.strip().lower(), "range": "0-999"}
        response = requests.get(f"{GLPI_URL}/search/User", headers=headers, params=params, verify=False)

        if response.status_code == 200:
            users = response.json().get("data", [])

            for user in users:
                # Manejar valores None de forma segura
                first_name = (user.get('9') or '').strip()
                last_name = (user.get('34') or '').strip()
                username_glpi = (user.get("1") or '').strip()
                
                # Construir el nombre completo
                nombre_completo = f"{first_name} {last_name}".strip()

                # Comparar el nombre normalizado
                if nombre_completo.lower() == username.strip().lower() or username_glpi.lower() == username.strip().lower():
                    messagebox.showinfo("Información", f"Usuario encontrado: {nombre_completo}, ID: {user.get('1')}")
                    return user.get("1")  # Asegúrate de que '1' es el ID correcto en tu sistema GLPI

            messagebox.showerror("Información", f"No se encontró el usuario '{username}' en GLPI.")
            return None
        else:
            messagebox.showerror("Error", f"Error al buscar el usuario en GLPI: {response.status_code}")
            return None

if __name__ == "__main__":
    root = tk.Tk()
    app = GLPIApp(root)
    root.mainloop()