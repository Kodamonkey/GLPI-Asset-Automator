# ----- Librerias ------

import tkinter as tk
from tkinter import messagebox, simpledialog, ttk, Toplevel, Label, Button
import pandas as pd
import cv2  # Para la captura de QR
from pyzbar.pyzbar import decode  # Decodificar QR
import os
import requests
import json
from dotenv import load_dotenv
import urllib3
import re
import numpy as np
import threading
import queue
from openpyxl import load_workbook, Workbook
from openpyxl.utils.dataframe import dataframe_to_rows

# ---------- Configuraciones -------------

# Deshabilitar las advertencias de SSL
urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)

# Cargar las variables del archivo .env
load_dotenv()

# Configuración de la API de GLPI
GLPI_URL = os.getenv("GLPI_URL")
USER_TOKEN = os.getenv("USER_TOKEN")
APP_TOKEN = os.getenv("APP_TOKEN")
PATH_EXCEL_ACTIVOS = os.getenv("PATH_EXCEL_ACTIVOS")
PATH_EXCEL_CONSUMIBLES = os.getenv("PATH_EXCEL_CONSUMIBLES")
IP_CAM_URL = os.getenv("IP_CAM_URL")

# Ruta del archivo Excel
ruta_excel = PATH_EXCEL_ACTIVOS

# Crear archivo Excel si no existe
def crear_archivo_excel_con_hojas(ruta, hojas):
    if not os.path.exists(ruta):
        wb = Workbook()
        for hoja in hojas:
            ws = wb.create_sheet(title=hoja)
            excel_headers = [
                "id", "asset_type", "entities_id", "name", "serial", "otherserial", "contact", "contact_num", 
                "users_id_tech", "groups_id_tech", "comment", "date_mod", "autoupdatesystems_id", 
                "locations_id", "networks_id", "computermodels_id", "computertypes_id", "is_template", 
                "template_name", "manufacturers_id", "is_deleted", "is_dynamic", "users_id", "groups_id", 
                "states_id", "ticket_tco", "uuid", "date_creation", "is_recursive","stock_target", "last_inventory_update", 
                "last_boot", "type", "model", "asset_tag", "purchase_date", "warranty_expiration_date", 
                "status", "location", "department", "ip_address", "mac_address", "operating_system", 
                "processor", "ram", "storage", "last_user", "supplier", "purchase_price", "order_number", 
                "invoice_number"
            ]
            ws.append(excel_headers)
        # Eliminar la hoja por defecto creada por Workbook
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        wb.save(ruta)

def crear_hoja_excel(wb, asset_type):
    if asset_type not in wb.sheetnames:
        ws = wb.create_sheet(title=asset_type)
        excel_headers = [
            "id", "asset_type", "entities_id", "name", "serial", "otherserial", "contact", "contact_num", 
            "users_id_tech", "groups_id_tech", "comment", "date_mod", "autoupdatesystems_id", 
            "locations_id", "networks_id", "computermodels_id", "computertypes_id", "is_template", 
            "template_name", "manufacturers_id", "is_deleted", "is_dynamic", "users_id", "groups_id", 
            "states_id", "ticket_tco", "uuid", "date_creation", "is_recursive","stock_target", "last_inventory_update", 
            "last_boot", "type", "model", "asset_tag", "purchase_date", "warranty_expiration_date", 
            "status", "location", "department", "ip_address", "mac_address", "operating_system", 
            "processor", "ram", "storage", "last_user", "supplier", "purchase_price", "order_number", 
            "invoice_number"
        ]
        ws.append(excel_headers)
    else:
        ws = wb[asset_type]
        excel_headers = [cell.value for cell in ws[1]]
    return ws, excel_headers

# Crear el archivo Excel con las hojas "Computer", "Monitor" y "Consumables" si no existe
crear_archivo_excel_con_hojas(ruta_excel, ["Computer", "Monitor", "Consumables"])
